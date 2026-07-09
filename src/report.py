# -*- coding: utf-8 -*-
"""
Формирование итогового Excel-отчёта (ТЗ, п.7): 8 листов —
Сводка / Сопоставленные / Расхождения / Не найдено у Покупателя /
Не найдено у Поставщика / Ручная проверка / Исходные данные / Журнал обработки.
"""
from __future__ import annotations
from typing import List
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import OperationCard
from .matching import MatchResult
from .metrics import Metrics

ALGO_VERSION = "reconcile-poc/1.0"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
NOTE_FONT = Font(name=FONT_NAME, italic=True, color="666666")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOOD_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
WARN_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
BAD_FILL = PatternFill("solid", start_color="FCE4E4", end_color="FCE4E4")


def _style_header(ws, row=1, ncols=1):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt_date(d):
    return d.strftime("%d.%m.%Y") if isinstance(d, datetime.date) else ""


def _write_rows(ws, start_row, rows):
    for r_i, row in enumerate(rows, start=start_row):
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = BORDER


def build_report(path: str, side_a: List[OperationCard], side_b: List[OperationCard],
                  results: List[MatchResult], missing_a: List[OperationCard],
                  missing_b: List[OperationCard], metrics: Metrics, config_name: str):
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, side_a, side_b, results, missing_a, missing_b, metrics, config_name)
    _sheet_matched(wb, results)
    _sheet_discrepancies(wb, results)
    _sheet_missing(wb, "Не найдено у Покупателя", missing_b,
                    "Документы есть у Стороны Б (Поставщик), но отсутствуют у Стороны А (Покупатель).")
    _sheet_missing(wb, "Не найдено у Поставщика", missing_a,
                    "Документы есть у Стороны А (Покупатель), но отсутствуют у Стороны Б (Поставщик).")
    _sheet_manual_review(wb, results)
    _sheet_raw_data(wb, side_a, side_b)
    _sheet_processing_log(wb, side_a, side_b, results)

    wb.save(path)


def _sheet_summary(wb, side_a, side_b, results, missing_a, missing_b, metrics, config_name):
    ws = wb.create_sheet("Сводка")
    ws["A1"] = "Сверка актов взаимных расчётов — сводка результатов"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Режим проверки: {config_name}. Версия алгоритма: {ALGO_VERSION}. Дата формирования: {datetime.date.today().strftime('%d.%m.%Y')}"
    ws["A2"].font = NOTE_FONT

    full_match = sum(1 for r in results if r.status == "Полное совпадение")
    other_match = sum(1 for r in results if r.status.startswith(("Совпадение", "Вероятное")))
    manual = sum(1 for r in results if "ручную" in r.status.lower() or "ручн" in r.status.lower() or "дубль" in r.status.lower())
    discrepancy = sum(1 for r in results if r.status.startswith("Расхождение"))
    f_level = sum(1 for r in results if r.level == "F")
    total_a, total_b = len(side_a), len(side_b)
    auto_matched_pct = round(100 * (full_match + other_match) / total_a, 1) if total_a else 0

    discrepancy_sum = sum(abs(r.amount_diff) for r in results if r.status.startswith("Расхождение"))
    missing_a_sum = sum(c.amount for c in missing_a)
    missing_b_sum = sum(c.amount for c in missing_b)

    rows = [
        ("Строк у Стороны А (Покупатель)", total_a),
        ("Строк у Стороны Б (Поставщик)", total_b),
        ("", ""),
        ("Точные совпадения (уровень A)", full_match),
        ("Вероятные/по реквизитам совпадения (уровни B/C/D/E)", other_match),
        ("Разбивка суммы на несколько строк (уровень F)", f_level),
        ("Расхождения по сумме/дате (номер совпал)", discrepancy),
        ("На ручную проверку (неоднозначные/дубли)", manual),
        ("Отсутствует у Покупателя (есть у Б, нет у А)", len(missing_b)),
        ("Отсутствует у Поставщика (есть у А, нет у Б)", len(missing_a)),
        ("", ""),
        ("Сумма расхождений по сумме/дате, руб.", round(discrepancy_sum, 2)),
        ("Сумма 'отсутствует у Покупателя', руб.", round(missing_b_sum, 2)),
        ("Сумма 'отсутствует у Поставщика', руб.", round(missing_a_sum, 2)),
        ("Доля автоматического сопоставления от строк Стороны А, %", auto_matched_pct),
        ("", ""),
        ("--- Сверка с эталоном (Итоги_для_проверки.xlsx) ---", ""),
        ("Эталонных расхождений", metrics.reference_count),
        ("Сумма эталона, руб.", metrics.reference_sum),
        ("Найдено системой (категория 'Не найдено у Покупателя')", metrics.predicted_count),
        ("Сумма найденного, руб.", metrics.predicted_sum),
        ("Найдено правильно (TP)", metrics.true_positive),
        ("Пропущено, ложноотрицательные (FN)", metrics.false_negative),
        ("Ложных срабатываний (FP)", metrics.false_positive),
        ("Точность (Precision), %", round(metrics.precision * 100, 1)),
        ("Полнота (Recall), %", round(metrics.recall * 100, 1)),
        ("F1-мера", metrics.f1),
        ("Совпадение суммы расхождений, %", round(100 * metrics.predicted_sum / metrics.reference_sum, 1) if metrics.reference_sum else 0),
    ]
    _write_rows(ws, 4, rows)
    for r in range(4, 4 + len(rows)):
        ws.cell(row=r, column=1).font = Font(name=FONT_NAME, size=10, bold=True)
    ws.cell(row=4 + len(rows) + 2, column=1,
            value="Примечание: методика расчёта метрик и известные ограничения источника данных — см. ТЕХНИЧЕСКИЙ_РАЗБОР.md.").font = NOTE_FONT
    _autofit(ws, [55, 20])


def _sheet_matched(wb, results):
    ws = wb.create_sheet("Сопоставленные")
    headers = ["Уровень", "Статус", "Уверенность, %", "Сторона А: документ", "Сторона А: номер",
               "Сторона А: дата", "Сторона А: сумма", "Сторона Б: документ", "Сторона Б: номер",
               "Сторона Б: дата", "Сторона Б: сумма", "Объяснение"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, ncols=len(headers))

    rows = []
    for r in results:
        if r.side_a is None or r.side_b is None:
            continue
        if not (r.level in ("A", "B", "C", "D", "E") or r.status.startswith("Расхождение")):
            continue
        a, b = r.side_a, r.side_b
        rows.append([r.level or "-", r.status, r.confidence, a.raw_document_text, a.document_number_raw,
                     _fmt_date(a.document_date), a.amount, b.raw_document_text, b.document_number_raw,
                     _fmt_date(b.document_date), b.amount, r.explanation])
    _write_rows(ws, 2, rows)
    for i, row in enumerate(rows, start=2):
        fill = GOOD_FILL if row[0] == "A" else (WARN_FILL if row[1].startswith("Расхождение") else None)
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=i, column=c).fill = fill
    _autofit(ws, [8, 34, 34, 14, 12, 34, 34, 14, 12, 45])
    ws.freeze_panes = "A2"


def _sheet_discrepancies(wb, results):
    ws = wb.create_sheet("Расхождения")
    headers = ["Статус", "Сторона А: документ", "Сторона А: сумма", "Сторона А: дата",
               "Сторона Б: документ", "Сторона Б: сумма", "Сторона Б: дата",
               "Δ Сумма", "Δ Дата, дн.", "Объяснение"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, ncols=len(headers))
    rows = []
    for r in results:
        if not r.status.startswith("Расхождение"):
            continue
        a, b = r.side_a, r.side_b
        rows.append([r.status, a.raw_document_text if a else "", a.amount if a else "",
                     _fmt_date(a.document_date) if a else "", b.raw_document_text if b else "",
                     b.amount if b else "", _fmt_date(b.document_date) if b else "",
                     r.amount_diff, r.date_diff_days, r.explanation])
    _write_rows(ws, 2, rows)
    for i in range(2, 2 + len(rows)):
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).fill = BAD_FILL
    _autofit(ws, [22, 34, 12, 14, 34, 12, 14, 10, 10, 45])
    ws.freeze_panes = "A2"


def _sheet_missing(wb, title, cards, note):
    ws = wb.create_sheet(title)
    ws["A1"] = note
    ws["A1"].font = NOTE_FONT
    headers = ["Документ (как напечатано)", "Тип", "Номер", "Дата документа", "Сумма", "Направление",
               "Источник (файл/страница/строка)", "Исходный текст строки"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, row=2, ncols=len(headers))
    rows = []
    for c in sorted(cards, key=lambda x: (x.document_date or datetime.date.min)):
        rows.append([c.raw_document_text, c.doc_type_raw, c.document_number_raw, _fmt_date(c.document_date),
                     c.amount, c.direction, f"{c.source_file} p{c.page} r{c.row}", c.source_text])
    _write_rows(ws, 3, rows)
    total_row = 3 + len(rows) + 1
    ws.cell(row=total_row, column=1, value="ИТОГО:").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=total_row, column=5, value=round(sum(c.amount for c in cards), 2)).font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
    _autofit(ws, [40, 14, 12, 14, 12, 12, 26, 50])


def _sheet_manual_review(wb, results):
    ws = wb.create_sheet("Ручная проверка")
    headers = ["Причина", "Сторона А", "Сторона Б", "Кол-во кандидатов", "Баллы/уверенность", "Рекомендация"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, ncols=len(headers))
    rows = []
    for r in results:
        is_manual = ("ручн" in r.status.lower()) or ("дубль" in r.status.lower()) or (r.level == "F")
        if not is_manual:
            continue
        a_desc = r.side_a.raw_document_text if r.side_a else (
            "; ".join(c.raw_document_text for c in (r.side_a_group or [])))
        b_desc = r.side_b.raw_document_text if r.side_b else (
            "; ".join(c.raw_document_text for c in (r.side_b_group or [])))
        rec = {
            "Неоднозначное совпадение — ручная проверка": "Проверить вручную по первичным документам — несколько кандидатов с близким баллом.",
            "Возможный дубль (Сторона А)": "Проверить, не введена ли операция дважды в учёте Стороны А.",
            "Возможный дубль (Сторона Б)": "Проверить, не введена ли операция дважды в учёте Стороны Б.",
        }.get(r.status, "Подтвердить разбивку суммы по первичным документам (акт может быть разнесён на несколько накладных).")
        rows.append([r.status, a_desc, b_desc, r.candidates_considered or "", r.confidence, rec])
    _write_rows(ws, 2, rows)
    _autofit(ws, [30, 40, 40, 16, 16, 50])
    ws.freeze_panes = "A2"


def _sheet_raw_data(wb, side_a, side_b):
    ws = wb.create_sheet("Исходные данные")
    headers = ["Сторона", "Файл", "№ строки", "Исходный текст строки (без изменений)",
               "Тип документа (распознан)", "Номер (сырой)", "Дата документа", "Дата проводки", "Сумма", "Направление"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, ncols=len(headers))
    rows = []
    for c in side_a + side_b:
        side_label = "А (Покупатель)" if c.side == "A" else "Б (Поставщик)"
        rows.append([side_label, c.source_file, c.row, c.source_text, c.doc_type_raw, c.document_number_raw,
                     _fmt_date(c.document_date), _fmt_date(c.accounting_date), c.amount, c.direction])
    _write_rows(ws, 2, rows)
    _autofit(ws, [14, 20, 8, 55, 16, 12, 14, 14, 12, 10])
    ws.freeze_panes = "A2"


def _sheet_processing_log(wb, side_a, side_b, results):
    ws = wb.create_sheet("Журнал обработки")
    headers = ["Файл", "Страница", "Строка", "OCR-уверенность", "Ошибки/примечания",
               "Применённое правило", "Версия алгоритма"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, ncols=len(headers))
    rule_by_key = {}
    for r in results:
        for c in [r.side_a, r.side_b] + (r.side_a_group or []) + (r.side_b_group or []):
            if c is not None:
                rule_by_key[c.card_id] = r.status
    rows = []
    for c in side_a + side_b:
        notes = "; ".join(c.processing_notes) if c.processing_notes else ""
        rule = rule_by_key.get(c.card_id, "не классифицировано (см. 'Не найдено у ...')")
        rows.append([c.source_file, c.page, c.row, c.ocr_confidence, notes, rule, ALGO_VERSION])
    _write_rows(ws, 2, rows)
    _autofit(ws, [20, 10, 8, 14, 45, 45, 16])
    ws.freeze_panes = "A2"
